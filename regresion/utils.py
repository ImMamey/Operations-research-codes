import openpyxl
import os
import array
import numpy as np
import pandas as pd


# os.chdir(r"/regresion/")
def load_data():
    workbook = openpyxl.load_workbook("DATOS PARA TRABAJO GRUPOS I.xlsx")
    sheet = workbook["GRUPO I"]

    "Array para las fechas y niveles de mareas de estacion el Tigre"
    M1_TR_FECHAS = []
    M1_TR_NIVEL_MAREAS = []
    M2_TR_FECHAS = []
    M2_TR_NIVEL_MAREAS = []

    for row in range(5, sheet.max_row + 1):
        "Se guarda iterativamente los valores de la columna C y D en su respectivo array"
        valor_fecha_1 = sheet.cell(row=row, column=3).value
        valor_marea_1 = sheet.cell(row=row, column=4).value
        valor_fecha_2 = sheet.cell(row=row, column=7).value
        valor_marea_2 = sheet.cell(row=row, column=8).value

        if (
            not valor_marea_1
            or not valor_fecha_1
            or not valor_marea_2
            or not valor_fecha_2
        ):
            "Si el valor de la celda es None, detiene el programa"
            break

        M1_TR_FECHAS.append(valor_fecha_1)
        M1_TR_NIVEL_MAREAS.append(valor_marea_1)
        M2_TR_FECHAS.append(valor_fecha_2)
        M2_TR_NIVEL_MAREAS.append(valor_marea_2)

    # TODO: la informacion de M1_TR_FECHAS es de tipo datetime, se debe convertir a string para la regresion lineal
    m1_fechas = np.array([dt.timestamp() for dt in M1_TR_FECHAS]).astype(int)
    m1_mareas = np.array(M1_TR_NIVEL_MAREAS)
    # 2_fechas = np.array([(date - M2_TR_FECHAS[0]).days for date in M2_TR_FECHAS])
    # m2_mareas = np.array(M2_TR_NIVEL_MAREAS)
    print("══════════════════════════════════════════")

    # x,y
    return m1_fechas, m1_mareas


# if __name__ == "__main__":
#    load_data()
